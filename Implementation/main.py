# Principles of Programming Languages and Compiler Design (CEID - University of Patras)
# Python Project 2021
# Christos-Panagiotis Balatsouras, StudentID: 1054335

# import python libraries
import requests
from openpyxl import load_workbook
from matplotlib.ticker import FuncFormatter
import matplotlib.pyplot as plt
import numpy as np
import warnings
import mysql.connector
import csv


# Global Variables
files_count = 0
Country_1 = "Greece"
Country_2 = "Sweden"
Years = [2016, 2017, 2018, 2019]
mydb = mysql.connector.connect(
  host="localhost",
  user="root",
  password="",
  database="1054335_python_project_2021"
)


class Dataset:
    def __init__(self, country1, country2, years):
        self.first_country = country1
        self.second_country = country2
        self.data1 = []
        self.data2 = []
        self.interval = years


def download(files_downloaded_count):  # download data in .xlsx format from the WEB
    # URLs
    url_nights_spent_residents = "https://ec.europa.eu/eurostat/databrowser-backend/api/query/1.0/LIVE/xlsx/en/download/24190e8b-f8c7-404c-960f-2c24fb85ca8c"
    url_nights_spent_non_residents = "https://ec.europa.eu/eurostat/databrowser-backend/api/query/1.0/LIVE/xlsx/en/download/1fdc52f3-abc1-413b-b1c9-3e8f93ab133a"
    url_arrivals_residents = "https://ec.europa.eu/eurostat/databrowser-backend/api/query/1.0/LIVE/xlsx/en/download/343a8149-aae7-46b5-b0d2-7e145484b060"
    url_arrivals_non_residents = "https://ec.europa.eu/eurostat/databrowser-backend/api/query/1.0/LIVE/xlsx/en/download/7ff88378-1f46-4644-82fd-09a35058e7df"

    # Download process
    print("Downloading: Nights spent at tourist accommodation establishments by residents, from URL: ", url_nights_spent_residents)
    req_nights_spent_residents = requests.get(url_nights_spent_residents, allow_redirects=True)
    open('nights_spent_by_residents.xlsx', 'wb').write(req_nights_spent_residents.content)
    print("File Downloaded Successfully")
    files_downloaded_count += 1

    print("Downloading: Nights spent at tourist accommodation establishments by non-residents, from URL: ", url_nights_spent_non_residents)
    req_nights_spent_non_residents = requests.get(url_nights_spent_non_residents, allow_redirects=True)
    open('nights_spent_by_non_residents.xlsx', 'wb').write(req_nights_spent_non_residents.content)
    print("File Downloaded Successfully")
    files_downloaded_count += 1

    print("Downloading: Arrivals of residents at tourist accommodation establishments, from URL: ", url_arrivals_residents)
    req_arrivals_residents = requests.get(url_arrivals_residents, allow_redirects=True)
    open('arrivals_by_residents.xlsx', 'wb').write(req_arrivals_residents.content)
    print("File Downloaded Successfully")
    files_downloaded_count += 1

    print("Downloading: Arrivals of non-residents at tourist accommodation establishments, from URL: ", url_arrivals_non_residents)
    req_arrivals_non_residents = requests.get(url_arrivals_non_residents, allow_redirects=True)
    open('arrivals_by_non_residents.xlsx', 'wb').write(req_arrivals_non_residents.content)
    print("File Downloaded Successfully")
    files_downloaded_count += 1

    print("\nDownload Completed Successfully!\nDownloaded ", files_downloaded_count, "files.\n")


def read_excel_data(file_name):  # Extract Useful Information from Excel Files
    dataset = Dataset(Country_1, Country_2, Years)
    year_columns = []
    row1 = None
    row2 = None

    with warnings.catch_warnings(record=True):
        warnings.simplefilter("always")
        wb = load_workbook(file_name)
    sheet_1 = wb['Sheet 1']
    for row in sheet_1.iter_rows(min_row=10, max_col=25, max_row=53):
        for cell in row:
            if cell.value == Country_1:
                row1 = cell.row
    for row in sheet_1.iter_rows(min_row=10, max_col=25, max_row=53):
        for cell in row:
            if cell.value == Country_2:
                row2 = cell.row
    for row in sheet_1.iter_rows(min_row=10, max_col=25, max_row=10):
        for cell in row:
            for item in Years:
                if cell.value == str(item):
                    col = cell.column
                    year_columns.append(col)

    numofyears = len(year_columns)
    # print(row1)
    # print(row2)
    # print(year_columns)

    for row in sheet_1.iter_rows(min_col=year_columns[0], min_row=row1, max_col=year_columns[numofyears-1], max_row=row1):
        for cell in row:
            if isinstance(cell.value, float):
                dataset.data1.append(int(cell.value))
    # print(dataset.data1)

    for row in sheet_1.iter_rows(min_col=year_columns[0], min_row=row2, max_col=year_columns[numofyears-1], max_row=row2):
        for cell in row:
            if isinstance(cell.value, float):
                dataset.data2.append(int(cell.value))
    # print(dataset.data2)

    return dataset


def format_millions(x, pos):  # Function for the formatting of the numbers in the plot
    return '%1.0fM' % (x * 1e-6)


def plot(country1, country2, years, data1, data2, xlabel, ylabel, title):
    x_pos = np.arange(len(years))
    width = 0.25  # bar width

    formatter = FuncFormatter(format_millions)

    fig, ax = plt.subplots()
    r1 = ax.bar(x_pos - width / 2, data1, width, label=country1, color='b')
    r2 = ax.bar(x_pos + width / 2, data2, width, label=country2, color='g')

    ax.set_ylabel(ylabel)
    ax.set_xlabel(xlabel)
    ax.set_title(title)
    ax.set_xticks(x_pos)
    ax.yaxis.set_major_formatter(formatter)
    ax.set_xticklabels(years)
    ax.legend()

    fig.tight_layout()
    plt.show()


def main():
    print("Principles of Programing Languages and Compiler Design")
    print("Python Project 2021")
    print("Author: Christos-Panagiotis Mpalatsouras, SudentID = 1054335\n")
    print("This script extracts information about tourism in Europe")
    print("First Country of Interest: ", Country_1)
    print("Second Country of Interest: ", Country_2)
    print("Years of interest: ", Years)

    print("Requested plots: ")
    print("1. Nights spent at tourist accommodation establishments")
    print("2. Nights spent by non-residents at tourist accommodation establishments")
    print("3. Arrivals at tourist accommodation establishments")
    print("4. Arrivals of non-residents at tourist accommodation establishments")

    print("\nThis Script Will Download Required Data from Eurostat Website")
    download(files_count)

    print("Extracting useful information from the downloaded files...")
    Nights_spent_residents = read_excel_data('nights_spent_by_residents.xlsx')
    Nights_spent_non_residents = read_excel_data('nights_spent_by_non_residents.xlsx')
    Arrivals_residents = read_excel_data('arrivals_by_residents.xlsx')
    Arrivals_non_residents = read_excel_data('arrivals_by_non_residents.xlsx')

    # Plotting the data
    print("Plotting the data...")
    plot(Nights_spent_residents.first_country, Nights_spent_residents.second_country, Years,
         Nights_spent_residents.data1, Nights_spent_residents.data2, "Year", "Million Nights Spent",
         "Nights spent at tourist accommodation establishments by residents")
    plot(Nights_spent_non_residents.first_country, Nights_spent_non_residents.second_country, Years,
         Nights_spent_non_residents.data1, Nights_spent_non_residents.data2, "Year", "Million Nights Spent",
         "Nights spent at tourist accommodation establishments by non-residents")
    plot(Arrivals_residents.first_country, Arrivals_residents.second_country, Years,
         Arrivals_residents.data1, Arrivals_residents.data2, "Year", "Million Arrivals",
         "Arrivals of residents at tourist accommodation establishments")
    plot(Arrivals_non_residents.first_country, Arrivals_non_residents.second_country, Years,
         Arrivals_non_residents.data1, Arrivals_non_residents.data2, "Year", "Million Arrivals",
         "Arrivals of non-residents at tourist accommodation establishments")

    # Writing data to Database
    mycursor = mydb.cursor()
    id = 1
    increment = 0
    print("Writing to database: Nights spent by residents")
    query = "INSERT INTO nights_spent (id, country, year, nights_spent) VALUES (%s, %s, %s, %s)"
    for item in Nights_spent_residents.data1:
        val = (id, Country_1, Nights_spent_residents.interval[increment], item)
        mycursor.execute(query, val)
        mydb.commit()
        increment += 1
        id += 1
    increment = 0
    for item in Nights_spent_residents.data2:
        val = (id, Country_2, Nights_spent_residents.interval[increment], item)
        mycursor.execute(query, val)
        mydb.commit()
        increment += 1
        id += 1
    increment = 0
    id = 1
    print("Writing to database: Nights spent by non residents")
    query = "INSERT INTO nights_spent_non_residents (id, country, year, nights_spent) VALUES (%s, %s, %s, %s)"
    for item in Nights_spent_non_residents.data1:
        val = (id, Country_1, Nights_spent_non_residents.interval[increment], item)
        mycursor.execute(query, val)
        mydb.commit()
        increment += 1
        id += 1
    increment = 0
    for item in Nights_spent_non_residents.data2:
        val = (id, Country_2, Nights_spent_non_residents.interval[increment], item)
        mycursor.execute(query, val)
        mydb.commit()
        increment += 1
        id += 1
    increment = 0
    id = 1
    print("Writing to database: Arrivals by residents")
    query = "INSERT INTO arrivals (id, country, year, arrivals) VALUES (%s, %s, %s, %s)"
    for item in Arrivals_residents.data1:
        val = (id, Country_1, Arrivals_residents.interval[increment], item)
        mycursor.execute(query, val)
        mydb.commit()
        increment += 1
        id += 1
    increment = 0
    for item in Arrivals_residents.data2:
        val = (id, Country_2, Arrivals_residents.interval[increment], item)
        mycursor.execute(query, val)
        mydb.commit()
        increment += 1
        id += 1
    increment = 0
    id = 1
    print("Writing to database: Arrivals by non residents")
    query = "INSERT INTO arrivals_non_residents (id, country, year, arrivals) VALUES (%s, %s, %s, %s)"
    for item in Arrivals_non_residents.data1:
        val = (id, Country_1, Arrivals_non_residents.interval[increment], item)
        mycursor.execute(query, val)
        mydb.commit()
        increment += 1
        id += 1
    increment = 0
    for item in Arrivals_non_residents.data2:
        val = (id, Country_2, Arrivals_non_residents.interval[increment], item)
        mycursor.execute(query, val)
        mydb.commit()
        increment += 1
        id += 1

    # Write CSV Files
    with open('nights_spent_residents_file.csv', mode='w', newline='') as nights_spent_residents_file:
        print("Extracting Nights Spent by Residents in CSV...")
        nights_spent_residents_writer = csv.writer(nights_spent_residents_file, delimiter=',', quotechar='"', quoting=csv.QUOTE_NONNUMERIC)
        index = 0
        nights_spent_residents_writer.writerow(['Country', 'Year', 'Nights_spent'])
        for item in Nights_spent_residents.data1:
            nights_spent_residents_writer.writerow([Country_1, Nights_spent_residents.interval[index], item])
            index += 1
        index = 0
        for item in Nights_spent_residents.data2:
            nights_spent_residents_writer.writerow([Country_2, Nights_spent_residents.interval[index], item])
            index += 1
    nights_spent_residents_file.close()

    with open('nights_spent_non_residents_file.csv', mode='w', newline='') as nights_spent_non_residents_file:
        print("Extracting Nights Spent by Non Residents in CSV...")
        nights_spent_non_residents_writer = csv.writer(nights_spent_non_residents_file, delimiter=',', quotechar='"', quoting=csv.QUOTE_NONNUMERIC)
        index = 0
        nights_spent_non_residents_writer.writerow(['Country', 'Year', 'Nights_spent'])
        for item in Nights_spent_non_residents.data1:
            nights_spent_non_residents_writer.writerow([Country_1, Nights_spent_non_residents.interval[index], item])
            index += 1
        index = 0
        for item in Nights_spent_non_residents.data2:
            nights_spent_non_residents_writer.writerow([Country_2, Nights_spent_non_residents.interval[index], item])
            index += 1
    nights_spent_non_residents_file.close()

    with open('arrivals_residents_file.csv', mode='w', newline='') as arrivals_residents_file:
        print("Extracting Arrivals by Residents in CSV...")
        arrivals_residents_writer = csv.writer(arrivals_residents_file, delimiter=',', quotechar='"', quoting=csv.QUOTE_NONNUMERIC)
        index = 0
        arrivals_residents_writer.writerow(['Country', 'Year', 'Arrivals'])
        for item in Arrivals_residents.data1:
            arrivals_residents_writer.writerow([Country_1, Arrivals_residents.interval[index], item])
            index += 1
        index = 0
        for item in Arrivals_residents.data2:
            arrivals_residents_writer.writerow([Country_2, Arrivals_residents.interval[index], item])
            index += 1
    arrivals_residents_file.close()

    with open('arrivals_non_residents_file.csv', mode='w', newline='') as arrivals_non_residents_file:
        print("Extracting Arrivals by Non Residents in CSV...")
        arrivals_non_residents_writer = csv.writer(arrivals_non_residents_file, delimiter=',', quotechar='"', quoting=csv.QUOTE_NONNUMERIC)
        index = 0
        arrivals_non_residents_writer.writerow(['Country', 'Year', 'Arrivals'])
        for item in Arrivals_non_residents.data1:
            arrivals_non_residents_writer.writerow([Country_1, Arrivals_non_residents.interval[index], item])
            index += 1
        index = 0
        for item in Arrivals_non_residents.data2:
            arrivals_non_residents_writer.writerow([Country_2, Arrivals_non_residents.interval[index], item])
            index += 1
    arrivals_non_residents_file.close()


main()
